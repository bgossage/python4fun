#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Apr  4 10:21:38 2018

@author: genecapture
"""

from openpyxl import load_workbook

myfilename = 'hybridization-profiles.xlsx'
mywb = load_workbook(filename = myfilename)
mysheet = mywb.get_sheet_by_name('All')
nr = mysheet.max_row
nc = mysheet.max_column

columns, rows = nc, nr;
mymatrix = [[0 for x in range(columns)] for y in range(rows)] 

for i in range(nr):
    for j in range(nc):
        mymatrix[i][j] = mysheet.cell(row=i+1,column=j+1).value

print (mymatrix[0][0], mymatrix[nr-2][nc-1])        

        
        